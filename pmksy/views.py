"""Application views for PMKSY import workflows."""

from __future__ import annotations

import csv
import logging
from collections import defaultdict, deque
from numbers import Integral
from pathlib import Path
from typing import Deque, Dict, Iterable, List, Optional, Tuple, Type

from django import forms
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import Http404, HttpRequest, HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from django.views import View

from data_wizard import InputNeeded, registry as data_wizard_registry
from data_wizard.models import Run
from data_wizard.sources.models import FileSource
from data_wizard.tasks import get_rows

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

from .importers import REGISTRATIONS
from .models import Farmer, ImportRunMetadata

logger = logging.getLogger(__name__)


PREVIEW_ERROR_MESSAGE = (
    "We couldn't read the uploaded file. Please upload a supported format."
)


class PreviewGenerationError(Exception):
    """Raised when a preview cannot be generated for an uploaded dataset."""


PREVIEW_LOAD_ERRORS: Tuple[Type[BaseException], ...] = (ValueError,)

try:  # pragma: no cover - optional dependency for XLS files
    from xlrd import XLRDError  # type: ignore
except ImportError:  # pragma: no cover - fallback when xlrd is unavailable
    XLRDError = None  # type: ignore[assignment]
else:  # pragma: no cover - exercised only when xlrd is installed
    PREVIEW_LOAD_ERRORS = PREVIEW_LOAD_ERRORS + (XLRDError,)

try:  # pragma: no cover - optional dependency for XLSX files
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:  # pragma: no cover - fallback when openpyxl is unavailable
    InvalidFileException = None  # type: ignore[assignment]
else:  # pragma: no cover - exercised only when openpyxl is installed
    PREVIEW_LOAD_ERRORS = PREVIEW_LOAD_ERRORS + (InvalidFileException,)

try:  # pragma: no cover - optional dependency used by data_wizard
    import itertable.exceptions as itertable_exceptions  # type: ignore
except ImportError:  # pragma: no cover - fallback when itertable is unavailable
    itertable_exceptions = None  # type: ignore[assignment]
else:  # pragma: no cover - exercised only when itertable is installed
    _itertable_error_types: List[Type[BaseException]] = []
    for name in ("TableError", "TableLoadError", "UnsupportedFormat"):
        exc = getattr(itertable_exceptions, name, None)
        if isinstance(exc, type) and issubclass(exc, BaseException):
            _itertable_error_types.append(exc)
    if _itertable_error_types:
        PREVIEW_LOAD_ERRORS = PREVIEW_LOAD_ERRORS + tuple(_itertable_error_types)


try:  # pragma: no cover - compatibility for future data_wizard releases
    from data_wizard.views import ImportWizard as BaseImportWizard  # type: ignore
except (ImportError, AttributeError):  # pragma: no cover - graceful fallback
    BaseImportWizard = View


class FileUploadForm(forms.Form):
    """Collect the upload, plus optional workbook sheet selection."""

    _WORKBOOK_EXTENSIONS = {".xls", ".xlsx"}

    source_file = forms.FileField(
        label="Upload data file",
        help_text="Supported formats: CSV, XLSX, XLS, JSON",
    )
    sheet_name = forms.ChoiceField(
        label="Sheet name",
        required=False,
        choices=(),
        widget=forms.Select,
        help_text=(
            "For Excel workbooks, provide the sheet to import. Leave blank for "
            "CSV or JSON files."
        ),
    )

    def clean_sheet_name(self) -> str:
        sheet_name = self.cleaned_data.get("sheet_name", "")
        if isinstance(sheet_name, str):
            return sheet_name.strip()
        return ""

    def clean(self) -> Dict[str, object]:
        cleaned_data = super().clean()
        uploaded_file = cleaned_data.get("source_file")
        sheet_name = cleaned_data.get("sheet_name")

        if uploaded_file and self._is_workbook(uploaded_file):
            if not sheet_name:
                self.add_error(
                    "sheet_name",
                    "Select the worksheet to import from your workbook.",
                )

        return cleaned_data

    def _is_workbook(self, uploaded_file: forms.FileField) -> bool:
        """Return True when the uploaded file appears to be a workbook."""

        content_type = getattr(uploaded_file, "content_type", "") or ""
        normalized_type = content_type.lower()

        if "spreadsheet" in normalized_type or "excel" in normalized_type:
            return True

        filename = getattr(uploaded_file, "name", "")
        extension = Path(filename).suffix.lower()
        return extension in self._WORKBOOK_EXTENSIONS


class ConfirmImportForm(forms.Form):
    run_id = forms.IntegerField(widget=forms.HiddenInput)


def build_wizard_registry() -> Dict[str, Dict[str, object]]:
    """Return a slug-indexed registry of configured importers."""

    registry: Dict[str, Dict[str, object]] = {}
    for name, serializer in REGISTRATIONS:
        slug = slugify(name)
        registry[slug] = {
            "name": name,
            "serializer": serializer,
        }
    return registry


WIZARD_MAP = build_wizard_registry()


def get_preview_rows(run: Run, limit: int = 5) -> Tuple[List[str], List[List[str]]]:
    """Generate a small preview of uploaded data for confirmation."""

    Loader = data_wizard_registry.get_loader(run.loader)
    loader = Loader(run)

    headers: List[str] = []
    lookup_keys: List[str] = []
    rows: List[List[str]] = []

    try:
        table = loader.load_iter()

        if hasattr(table, "field_map") and table.field_map:
            headers = list(table.field_map.keys())

            def _normalize_lookup_key(key: object) -> object:
                if isinstance(key, Integral) and not isinstance(key, bool):
                    return str(key)
                return key

            lookup_keys = [_normalize_lookup_key(table.field_map[h]) for h in headers]

        iterator: Iterable = table
        for index, row in enumerate(iterator):
            if hasattr(row, "_asdict"):
                data = row._asdict()
            elif isinstance(row, dict):
                data = row
            elif isinstance(row, (list, tuple)):
                data = {str(i): value for i, value in enumerate(row)}
            else:
                data = {"value": row}

            if not headers:
                headers = list(data.keys())
                lookup_keys = list(headers)

            row_values: List[str] = []
            for column_index, column in enumerate(headers):
                lookup_key = (
                    lookup_keys[column_index]
                    if column_index < len(lookup_keys)
                    else column
                )
                value = data.get(lookup_key, data.get(column, ""))
                row_values.append(str(value))

            rows.append(row_values)

            if index + 1 >= limit:
                break
    except PREVIEW_LOAD_ERRORS as exc:
        raise PreviewGenerationError("Unable to load preview data") from exc

    return headers, rows


class RunAccessMixin:
    """Provide helpers for retrieving data wizard runs owned by the user."""

    def _get_user_run(self, run_id: int) -> Run:
        try:
            run = Run.objects.get(pk=run_id, user=self.request.user)
        except Run.DoesNotExist as exc:  # pragma: no cover - defensive
            raise Http404("Import session not found") from exc
        return run


class ImportHomeView(LoginRequiredMixin, View):
    template_name = "pmksy/home.html"

    def get(self, request: HttpRequest) -> HttpResponse:
        wizards = [
            {
                "name": config["name"],
                "slug": slug,
            }
            for slug, config in WIZARD_MAP.items()
        ]
        context = {"wizards": wizards}
        return render(request, self.template_name, context)


class PMKSYImportWizard(LoginRequiredMixin, RunAccessMixin, BaseImportWizard):
    upload_template_name = "pmksy/import_wizard_upload.html"
    preview_template_name = "pmksy/import_wizard_preview.html"
    success_template_name = "pmksy/wizard_done.html"

    form_class = FileUploadForm
    confirm_form_class = ConfirmImportForm

    def dispatch(self, request: HttpRequest, *args, **kwargs) -> HttpResponse:
        self.wizard_slug = kwargs.get("wizard_slug")
        if self.wizard_slug not in WIZARD_MAP:
            raise Http404("Unknown import wizard")
        self.wizard_config = WIZARD_MAP[self.wizard_slug]
        self._expected_fields_cache = None
        return super().dispatch(request, *args, **kwargs)

    def get(self, request: HttpRequest, *args, **kwargs) -> HttpResponse:
        run_id = request.GET.get("run")
        if run_id:
            run = self._get_user_run(run_id)
            try:
                headers, rows = get_preview_rows(run)
            except PreviewGenerationError:
                logger.exception("Failed to generate preview for run %s", run.pk)
                messages.error(request, PREVIEW_ERROR_MESSAGE)
                return redirect(
                    reverse("pmksy:wizard", kwargs={"wizard_slug": self.wizard_slug})
                )
            context = {
                "wizard": self.wizard_config,
                "wizard_slug": self.wizard_slug,
                "run": run,
                "headers": headers,
                "rows": rows,
                "confirm_form": self.confirm_form_class(initial={"run_id": run.pk}),
            }
            return render(request, self.preview_template_name, context)

        context = {
            "wizard": self.wizard_config,
            "wizard_slug": self.wizard_slug,
            "form": self.form_class(),
            "expected_fields": self._get_expected_fields(),
        }
        return render(request, self.upload_template_name, context)

    def post(self, request: HttpRequest, *args, **kwargs) -> HttpResponse:
        if "run_id" in request.POST:
            return self._handle_confirmation(request)
        return self._handle_upload(request)

    def _handle_upload(self, request: HttpRequest) -> HttpResponse:
        form = self.form_class(request.POST, request.FILES)
        uploaded_file = request.FILES.get("source_file")
        sheet_choices: List[Tuple[str, str]] = []

        if uploaded_file and form._is_workbook(uploaded_file):
            sheet_names = self._collect_workbook_sheetnames(uploaded_file)
            if sheet_names:
                sheet_choices = [(name, name) for name in sheet_names]

        if not sheet_choices:
            raw_sheet = request.POST.get("sheet_name")
            if isinstance(raw_sheet, str) and raw_sheet:
                display_label = raw_sheet.strip() or raw_sheet
                sheet_choices = [(raw_sheet, display_label)]

        if sheet_choices:
            form.fields["sheet_name"].choices = sheet_choices

        if not form.is_valid():
            if sheet_choices:
                form.fields["sheet_name"].choices = sheet_choices
            context = {
                "wizard": self.wizard_config,
                "wizard_slug": self.wizard_slug,
                "form": form,
                "expected_fields": self._get_expected_fields(),
            }
            return render(request, self.upload_template_name, context)

        uploaded_file = form.cleaned_data["source_file"]
        file_source = FileSource.objects.create(
            user=request.user,
            file=uploaded_file,
            name=f"{self.wizard_config['name']} Upload",
        )

        run = Run.objects.create(
            user=request.user,
            content_object=file_source,
            serializer=self.wizard_config["name"],
        )

        sheet_name = form.cleaned_data.get("sheet_name", "")
        if sheet_name:
            ImportRunMetadata.objects.update_or_create(
                run=run,
                defaults={"sheet_name": sheet_name},
            )
            self._warn_for_merged_cells(request, file_source, sheet_name)

        logger.info("Created run %s for wizard %s", run.pk, self.wizard_slug)

        return redirect(
            f"{reverse('pmksy:wizard', kwargs={'wizard_slug': self.wizard_slug})}?run={run.pk}"
        )

    def _collect_workbook_sheetnames(self, uploaded_file) -> List[str]:
        """Return the list of sheet names for the uploaded workbook."""

        workbook = None
        sheet_names: List[str] = []

        try:
            if hasattr(uploaded_file, "seek"):
                uploaded_file.seek(0)
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
            sheet_names = list(getattr(workbook, "sheetnames", []))
        except Exception:  # pragma: no cover - defensive guard against bad files
            logger.debug("Unable to enumerate workbook sheets", exc_info=True)
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:  # pragma: no cover - best-effort cleanup
                    logger.debug("Failed to close workbook after inspection", exc_info=True)
            if hasattr(uploaded_file, "seek"):
                try:
                    uploaded_file.seek(0)
                except Exception:  # pragma: no cover - best-effort reset
                    logger.debug("Unable to reset uploaded workbook pointer", exc_info=True)

        return sheet_names

    def _warn_for_merged_cells(
        self,
        request: HttpRequest,
        file_source: FileSource,
        sheet_name: str,
    ) -> None:
        """Surface a warning when the selected sheet includes merged cells."""

        file_field = getattr(file_source, "file", None)
        if not file_field:
            return

        workbook_source = getattr(file_field, "path", None)
        should_close_file = False

        if not workbook_source:
            try:
                file_field.open("rb")
            except Exception:  # pragma: no cover - storage backend edge case
                logger.warning(
                    "Unable to open uploaded workbook for merged-cell inspection",
                    exc_info=True,
                )
                return
            workbook_source = file_field
            should_close_file = True

        try:
            workbook = load_workbook(
                workbook_source,
                data_only=True,
            )
        except Exception:  # pragma: no cover - defensive guard against bad files
            logger.warning(
                "Unable to read workbook when checking for merged cells",
                exc_info=True,
            )
            if should_close_file:
                file_field.close()
            return

        try:
            try:
                sheet = workbook[sheet_name]
            except KeyError:
                logger.warning(
                    "Sheet '%s' missing while checking for merged cells", sheet_name
                )
                return

            merged_ranges: List[str] = []
            for merged_range in getattr(sheet.merged_cells, "ranges", []):
                coord = getattr(merged_range, "coord", None)
                if not coord:
                    bounds = getattr(merged_range, "bounds", None)
                    if bounds:
                        min_col, min_row, max_col, max_row = bounds
                        coord = (
                            f"{get_column_letter(min_col)}{min_row}:"
                            f"{get_column_letter(max_col)}{max_row}"
                        )
                merged_ranges.append(coord or str(merged_range))

            if merged_ranges:
                readable_ranges = ", ".join(merged_ranges)
                messages.warning(
                    request,
                    (
                        f"Sheet '{sheet_name}' contains merged cells: {readable_ranges}. "
                        "Please unmerge these cells before importing."
                    ),
                )
        finally:
            workbook.close()
            if should_close_file:
                file_field.close()

    def _handle_confirmation(self, request: HttpRequest) -> HttpResponse:
        form = self.confirm_form_class(request.POST)
        if not form.is_valid():
            messages.error(request, "Unable to confirm import. Please try again.")
            return redirect(
                reverse("pmksy:wizard", kwargs={"wizard_slug": self.wizard_slug})
            )

        run = self._get_user_run(form.cleaned_data["run_id"])

        try:
            result = run.run_all(run.get_auto_import_tasks())
        except InputNeeded as exc:  # pragma: no cover - defensive
            logger.warning("Import requires additional input: action=%s", exc.action)
            messages.warning(
                request,
                "Additional configuration is required before importing this dataset.",
            )
            return redirect(run.get_absolute_url())

        if isinstance(result, dict):
            if result.get("error"):
                messages.error(request, result["error"])
                return redirect(
                    f"{reverse('pmksy:wizard', kwargs={'wizard_slug': self.wizard_slug})}?run={run.pk}"
                )
            if result.get("action"):
                messages.warning(
                    request,
                    "Further setup is required via the advanced wizard interface.",
                )
                return redirect(run.get_absolute_url())

        messages.success(
            request,
            f"Successfully imported data for {self.wizard_config['name']}.",
        )

        run.refresh_from_db()
        success_count = run.record_set.filter(success=True).count()
        failed_records = run.record_set.filter(success=False)
        failure_count = failed_records.count()
        failure_preview = list(
            failed_records.values("row", "fail_reason")[:5]
        )

        processed_count = run.record_count
        if processed_count is None:
            processed_count = success_count + failure_count
        skipped_count = failure_count

        context = {
            "wizard": self.wizard_config,
            "wizard_slug": self.wizard_slug,
            "run": run,
            "processed_count": processed_count,
            "skipped_count": skipped_count,
            "success_count": success_count,
            "failure_count": failure_count,
            "failure_preview": failure_preview,
        }
        return render(request, self.success_template_name, context)

    def _get_expected_fields(self) -> List[Dict[str, object]]:
        cache = getattr(self, "_expected_fields_cache", None)
        if cache is not None:
            return cache

        serializer_class = self.wizard_config.get("serializer")
        expected_fields: List[Dict[str, object]] = []

        if serializer_class:
            serializer = serializer_class(context={"request": getattr(self, "request", None)})
            for name, field in serializer.get_fields().items():
                if getattr(field, "read_only", False):
                    continue

                model_field = getattr(field, "model_field", None)
                if model_field is not None and getattr(model_field, "auto_created", False):
                    continue

                label = field.label or name.replace("_", " ").title()
                help_text = getattr(field, "help_text", "") or ""
                expected_fields.append(
                    {
                        "name": name,
                        "label": label,
                        "required": getattr(field, "required", False),
                        "help_text": help_text,
                    }
                )

        self._expected_fields_cache = expected_fields
        return expected_fields


class FarmerRegistrationDownloadView(LoginRequiredMixin, RunAccessMixin, View):
    """Stream farmer registration IDs for a completed import run."""

    CSV_HEADERS = ["registration_id", "name", "village", "district", "contact_no"]

    def get(self, request: HttpRequest, run_id: int, *args, **kwargs) -> HttpResponse:
        run = self._get_user_run(run_id)

        if run.serializer != "Farmers":
            raise Http404("Farmer registration data not available for this import.")

        success_records = list(
            run.record_set.filter(success=True)
            .select_related("pmksy_label")
            .order_by("row")
        )

        if not success_records:
            raise Http404("No farmer registrations available for download.")

        table = run.load_iter()
        start_row = getattr(table, "start_row", 0)

        row_mapping: Dict[int, Dict[str, object]] = {}
        for index, row_data in enumerate(get_rows(run)):
            row_number = index + start_row
            if isinstance(row_data, dict):
                row_mapping[row_number] = row_data

        import_start: Optional[object] = (
            run.log.filter(event="do_import").values_list("date", flat=True).first()
        )
        import_end: Optional[object] = (
            run.log.filter(event="import_complete").values_list("date", flat=True).last()
        )
        if import_end is None:
            import_end = timezone.now()

        new_farmer_filters = {}
        if import_start is not None:
            new_farmer_filters["created_at__gte"] = import_start
        if import_end is not None:
            new_farmer_filters["created_at__lte"] = import_end

        if new_farmer_filters:
            new_farmers = list(
                Farmer.objects.filter(**new_farmer_filters).order_by("created_at", "farmer_id")
            )
        else:
            new_farmers = []

        new_farmers_by_name: Dict[str, Deque[Farmer]] = defaultdict(deque)
        for farmer in new_farmers:
            new_farmers_by_name[self._normalize_value(farmer.name)].append(farmer)

        unused_new_farmers: Deque[Farmer] = deque(new_farmers)
        used_farmer_ids = set()
        rows: List[List[str]] = []

        for record in success_records:
            row_data = row_mapping.get(record.row, {})
            registration_value = self._normalize_value(row_data.get("registration_id"))

            farmer: Optional[Farmer] = None
            if registration_value:
                farmer = Farmer.objects.filter(registration_id=registration_value).first()

            if farmer is None:
                label_source = row_data.get("name") if isinstance(row_data, dict) else None
                if not label_source and getattr(record, "pmksy_label", None) is not None:
                    label_source = record.pmksy_label.label
                name_value = self._normalize_value(label_source)

                if name_value:
                    queue = new_farmers_by_name.get(name_value)
                    while queue:
                        candidate = queue.popleft()
                        if candidate.pk not in used_farmer_ids:
                            farmer = candidate
                            break

                if farmer is None and unused_new_farmers:
                    while unused_new_farmers:
                        candidate = unused_new_farmers.popleft()
                        if candidate.pk not in used_farmer_ids:
                            farmer = candidate
                            break

                if farmer is None and name_value:
                    candidates = (
                        Farmer.objects.filter(name=name_value)
                        .order_by("created_at", "farmer_id")
                    )
                    for candidate in candidates:
                        if candidate.pk not in used_farmer_ids:
                            farmer = candidate
                            break

            if farmer is None:
                continue

            used_farmer_ids.add(farmer.pk)
            rows.append(
                [
                    self._normalize_value(farmer.registration_id),
                    self._normalize_value(farmer.name),
                    self._normalize_value(farmer.village),
                    self._normalize_value(farmer.district),
                    self._normalize_value(farmer.contact_no),
                ]
            )

        response = HttpResponse(content_type="text/csv")
        filename = f"farmers-run-{run.pk}-registrations.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow(self.CSV_HEADERS)
        for row in rows:
            writer.writerow(row)
        return response

    @staticmethod
    def _normalize_value(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()
