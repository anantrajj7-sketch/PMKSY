import csv
import io
import tempfile
from unittest import mock
from urllib.parse import parse_qs, urlparse

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse

from openpyxl import Workbook

from data_wizard.models import Run

from ..views import PREVIEW_ERROR_MESSAGE
from ..models import Farmer, ImportRunMetadata


class PMKSYImportWizardPreviewTests(TestCase):
    """Tests for PMKSY import wizard preview handling."""

    def setUp(self) -> None:
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="tester",
            email="tester@example.com",
            password="password123",
        )
        self.client.force_login(self.user)
        self.wizard_url = reverse("pmksy:wizard", kwargs={"wizard_slug": "farmers"})

    def test_invalid_workbook_redirects_with_message(self) -> None:
        """An invalid workbook should surface a helpful error message."""

        class BrokenLoader:
            def __init__(self, run):  # pragma: no cover - simple data holder
                self.run = run

            def load_iter(self):
                raise ValueError("Unsupported workbook format")

        uploaded_file = SimpleUploadedFile(
            "invalid.xlsx",
            b"not a real workbook",
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

        with mock.patch(
            "pmksy.views.data_wizard_registry.get_loader", return_value=BrokenLoader
        ):
            with self.assertLogs("pmksy.views", level="ERROR") as logs:
                response = self.client.post(
                    self.wizard_url,
                    {"source_file": uploaded_file, "sheet_name": "Data"},
                    follow=True,
                )

        self.assertEqual(response.status_code, 200)
        self.assertGreaterEqual(len(response.redirect_chain), 2)
        self.assertEqual(response.redirect_chain[-1][0], self.wizard_url)

        messages = [message.message for message in response.context["messages"]]
        self.assertIn(PREVIEW_ERROR_MESSAGE, messages)

        self.assertTrue(
            any("Failed to generate preview" in message for message in logs.output)
        )

    def test_upload_page_displays_field_guidance(self) -> None:
        """Users should see guidance describing the expected import columns."""

        response = self.client.get(
            reverse("pmksy:wizard", kwargs={"wizard_slug": "land-holdings"})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "<code>farmer</code>", html=True)
        self.assertContains(response, "Farmer Registration ID")

    def test_workbook_upload_requires_sheet(self) -> None:
        """Excel uploads must include the sheet selector."""

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetOne"
        sheet.append(["name"])
        sheet.append(["Row 1"])

        second_sheet = workbook.create_sheet(title="SheetTwo")
        second_sheet.append(["name"])
        second_sheet.append(["Row 2"])

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        uploaded_file = SimpleUploadedFile(
            "missing-sheet.xlsx",
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

        response = self.client.post(
            self.wizard_url,
            {"source_file": uploaded_file, "sheet_name": ""},
        )

        self.assertEqual(response.status_code, 200)
        form = response.context["form"]
        self.assertIn("sheet_name", form.errors)
        self.assertIn(
            "Select the worksheet to import from your workbook.",
            form.errors["sheet_name"],
        )

        choices = form.fields["sheet_name"].choices
        self.assertEqual(
            choices,
            [("SheetOne", "SheetOne"), ("SheetTwo", "SheetTwo")],
        )

        content = response.content.decode()
        self.assertInHTML(
            '<option value="" selected>Select a worksheet</option>', content
        )
        self.assertInHTML('<option value="SheetOne">SheetOne</option>', content)
        self.assertInHTML('<option value="SheetTwo">SheetTwo</option>', content)

        selected_sheet = choices[0][0]
        existing_run_ids = list(Run.objects.values_list("pk", flat=True))

        follow_up_buffer = io.BytesIO()
        workbook.save(follow_up_buffer)
        follow_up_buffer.seek(0)

        follow_up_file = SimpleUploadedFile(
            "selected-sheet.xlsx",
            follow_up_buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

        follow_up_response = self.client.post(
            self.wizard_url,
            {"source_file": follow_up_file, "sheet_name": selected_sheet},
            follow=True,
        )

        self.assertEqual(follow_up_response.status_code, 200)

        new_runs = Run.objects.exclude(pk__in=existing_run_ids)
        self.assertEqual(new_runs.count(), 1)
        run = new_runs.get()
        metadata = ImportRunMetadata.objects.get(run=run)
        self.assertEqual(metadata.sheet_name, selected_sheet)

    def test_selected_sheet_populates_preview(self) -> None:
        """The preview should reflect the user-selected worksheet."""

        workbook = Workbook()
        first_sheet = workbook.active
        first_sheet.title = "Overview"
        first_sheet.append(["name", "value"])
        first_sheet.append(["ignored", "0"])

        second_sheet = workbook.create_sheet(title="SurveyData")
        second_sheet.append(["name", "value"])
        second_sheet.append(["target", "42"])

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        uploaded_file = SimpleUploadedFile(
            "survey.xlsx",
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )

        response = self.client.post(
            self.wizard_url,
            {"source_file": uploaded_file, "sheet_name": "SurveyData"},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        rows = response.context["rows"]
        self.assertTrue(any("target" in row for row in rows))
        self.assertFalse(any("ignored" in row for row in rows))

        run = Run.objects.latest("pk")
        metadata = ImportRunMetadata.objects.get(run=run)
        self.assertEqual(metadata.sheet_name, "SurveyData")


class AuthenticationFlowTests(TestCase):
    """Ensure unauthenticated users are prompted to log in before importing."""

    def setUp(self) -> None:
        user_model = get_user_model()
        self.password = "password123"
        self.user = user_model.objects.create_user(
            username="authuser",
            email="authuser@example.com",
            password=self.password,
        )
        self.import_home_url = reverse("pmksy:import-home")
        self.login_url = reverse("login")

    def test_import_home_redirects_to_login(self) -> None:
        """Anonymous users should be redirected to the login page."""

        response = self.client.get(self.import_home_url)

        self.assertRedirects(
            response,
            f"{self.login_url}?next={self.import_home_url}",
            fetch_redirect_response=False,
        )

        login_page = self.client.get(response.url)
        self.assertEqual(login_page.status_code, 200)
        self.assertContains(login_page, "Sign in")

    def test_authenticated_user_can_access_import_home(self) -> None:
        """After logging in, users should reach the import home page."""

        logged_in = self.client.login(username=self.user.username, password=self.password)
        self.assertTrue(logged_in)

        response = self.client.get(self.import_home_url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "PMKSY Data Import")


class FarmerRegistrationDownloadTests(TestCase):
    """Verify the farmer registration download endpoint exposes imported IDs."""

    def setUp(self) -> None:
        user_model = get_user_model()
        self.user = user_model.objects.create_user(
            username="registrationtester",
            email="registrationtester@example.com",
            password="password123",
        )
        self.client.force_login(self.user)
        self.wizard_url = reverse("pmksy:wizard", kwargs={"wizard_slug": "farmers"})

    def test_download_includes_imported_registration_ids(self) -> None:
        """After importing farmers, users can download registration identifiers."""

        csv_payload = """name,village,district,contact_no
Alpha Farmer,Example Village,Example District,1111111111
Beta Farmer,Another Village,Another District,2222222222
"""

        with tempfile.TemporaryDirectory() as media_root, override_settings(
            MEDIA_ROOT=media_root
        ):
            upload = SimpleUploadedFile(
                "farmers.csv",
                csv_payload.encode("utf-8"),
                content_type="text/csv",
            )
            response = self.client.post(self.wizard_url, {"source_file": upload})

            self.assertEqual(response.status_code, 302)
            preview_url = response["Location"]
            run_id = int(parse_qs(urlparse(preview_url).query)["run"][0])

            preview_response = self.client.get(preview_url)
            self.assertEqual(preview_response.status_code, 200)

            confirm_response = self.client.post(
                self.wizard_url, {"run_id": run_id}, follow=True
            )
            self.assertEqual(confirm_response.status_code, 200)

            download_url = reverse(
                "pmksy:farmer-registration-download", kwargs={"run_id": run_id}
            )
            download_response = self.client.get(download_url)

        self.assertEqual(download_response.status_code, 200)
        self.assertEqual(download_response["Content-Type"], "text/csv")
        self.assertIn(
            f"farmers-run-{run_id}", download_response["Content-Disposition"]
        )

        reader = csv.DictReader(download_response.content.decode("utf-8").splitlines())
        self.assertEqual(
            reader.fieldnames,
            ["registration_id", "name", "village", "district", "contact_no"],
        )

        rows = list(reader)
        self.assertEqual(len(rows), 2)
        self.assertTrue(all(row["registration_id"] for row in rows))

        expected_records = {
            (
                farmer.registration_id,
                farmer.name,
                farmer.village,
                farmer.district,
                farmer.contact_no,
            )
            for farmer in Farmer.objects.order_by("created_at", "registration_id")
        }
        csv_records = {
            (
                row["registration_id"],
                row["name"],
                row["village"],
                row["district"],
                row["contact_no"],
            )
            for row in rows
        }
        self.assertSetEqual(csv_records, expected_records)
