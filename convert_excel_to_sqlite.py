"""Utility for converting Excel workbooks into a SQLite database.

This script is designed for messy survey-style spreadsheets where:

* Column headers might vary from file to file or worksheet to worksheet.
* Cells are sometimes merged (both across columns and across rows).
* There may be leading blank rows or descriptive headers before the
  tabular data actually starts.

Usage
-----

```bash
python convert_excel_to_sqlite.py --db output.sqlite data/*.xlsx
```

The script will create (or overwrite) ``output.sqlite`` and then create
one table per worksheet.  The table name is based on the file stem and
worksheet title (sanitised so that it is a valid SQLite identifier).
Original header names are preserved in a helper metadata table so that
you can refer back to the source labels.
"""

from __future__ import annotations

import argparse
import itertools
import re
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Iterator, List, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class WorksheetPayload:
    """Represents a worksheet that has been cleaned and converted."""

    table_name: str
    original_headers: Sequence[str]
    dataframe: pd.DataFrame
    source_file: Path
    sheet_name: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "excel_files",
        nargs="+",
        type=Path,
        help="Paths or glob patterns for the Excel workbooks to ingest.",
    )
    parser.add_argument(
        "--db",
        type=Path,
        default=Path("excel_data.sqlite"),
        help="Path to the SQLite database file to create or update.",
    )
    parser.add_argument(
        "--max-header-scan",
        type=int,
        default=10,
        help="Number of leading rows to scan when trying to detect the header row.",
    )
    return parser.parse_args()


def resolve_files(paths: Sequence[Path]) -> List[Path]:
    resolved: List[Path] = []
    for path in paths:
        if any(ch in str(path) for ch in "*?[]"):
            resolved.extend(sorted(Path().glob(str(path))))
        elif path.exists():
            resolved.append(path)
        else:
            raise FileNotFoundError(f"No such file or pattern: {path}")
    return resolved


def unmerge_cells(ws: Worksheet) -> None:
    """Replace merged cells with their top-left value."""

    for merged_range in list(ws.merged_cells.ranges):
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        )
        value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value in (None, ""):
                    cell.value = value


def iter_worksheet_values(ws: Worksheet) -> Iterator[Tuple]:
    return ws.values  # type: ignore[return-value]


def drop_leading_blank_rows(rows: List[Tuple]) -> List[Tuple]:
    iterator = iter(rows)
    remainder = list(itertools.dropwhile(_is_blank_row, iterator))
    return remainder


def _is_blank_row(row: Sequence) -> bool:
    return all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row)


def detect_header_index(rows: Sequence[Tuple], max_scan: int) -> int:
    max_scan = min(max_scan, len(rows))
    best_idx, best_score = 0, -1
    for idx in range(max_scan):
        row = rows[idx]
        score = 0
        for cell in row:
            if cell is None:
                continue
            if isinstance(cell, str):
                score += 1 if cell.strip() else 0
            else:
                score += 1
        if score > best_score:
            best_idx, best_score = idx, score
    return best_idx


def sanitise_identifier(text: str, fallback_prefix: str = "column") -> str:
    text = text.strip().lower()
    text = re.sub(r"[^0-9a-zA-Z_]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    if not text:
        text = fallback_prefix
    if text[0].isdigit():
        text = f"_{text}"
    return text


def build_headers(raw_headers: Sequence, fallback_prefix: str = "column") -> List[str]:
    counts: dict[str, int] = {}
    headers: List[str] = []
    for idx, header in enumerate(raw_headers):
        header_text = str(header) if header not in (None, "") else f"{fallback_prefix}_{idx + 1}"
        base_name = sanitise_identifier(header_text, fallback_prefix=f"{fallback_prefix}_{idx + 1}")
        count = counts.get(base_name, 0)
        counts[base_name] = count + 1
        if count:
            headers.append(f"{base_name}_{count + 1}")
        else:
            headers.append(base_name)
    return headers


def build_table_name(file_stem: str, sheet_name: str) -> str:
    base = f"{file_stem}_{sheet_name}"
    table_name = sanitise_identifier(base, fallback_prefix="table")
    return table_name[:63]  # SQLite accepts up to 64 characters comfortably.


def worksheet_to_dataframe(
    source_file: Path,
    ws: Worksheet,
    max_header_scan: int,
) -> WorksheetPayload | None:
    unmerge_cells(ws)
    rows = list(iter_worksheet_values(ws))
    rows = drop_leading_blank_rows(rows)
    if not rows:
        return None

    header_idx = detect_header_index(rows, max_header_scan)
    header_row = rows[header_idx]
    data_rows = rows[header_idx + 1 :]

    headers = build_headers(header_row)
    df = pd.DataFrame(data_rows, columns=headers)
    df = df.dropna(how="all")

    if df.empty:
        return None

    table_name = build_table_name(source_file.stem, ws.title)
    return WorksheetPayload(
        table_name=table_name,
        original_headers=[str(col) if col is not None else "" for col in header_row],
        dataframe=df,
        source_file=source_file,
        sheet_name=ws.title,
    )


def write_payloads_to_db(payloads: Iterable[WorksheetPayload], db_path: Path) -> None:
    conn = sqlite3.connect(db_path)
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS table_metadata (
                table_name TEXT PRIMARY KEY,
                source_file TEXT NOT NULL,
                worksheet TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS column_metadata (
                table_name TEXT NOT NULL,
                column_position INTEGER NOT NULL,
                column_name TEXT NOT NULL,
                original_header TEXT,
                PRIMARY KEY (table_name, column_position)
            )
            """
        )

        for payload in payloads:
            payload.dataframe.to_sql(payload.table_name, conn, if_exists="replace", index=False)
            conn.execute(
                "REPLACE INTO table_metadata (table_name, source_file, worksheet) VALUES (?, ?, ?)",
                (payload.table_name, str(payload.source_file), payload.sheet_name),
            )
            conn.executemany(
                "REPLACE INTO column_metadata (table_name, column_position, column_name, original_header) VALUES (?, ?, ?, ?)",
                [
                    (
                        payload.table_name,
                        idx,
                        payload.dataframe.columns[idx],
                        payload.original_headers[idx],
                    )
                    for idx in range(len(payload.dataframe.columns))
                ],
            )
        conn.commit()
    finally:
        conn.close()


def process_workbook(path: Path, max_header_scan: int) -> List[WorksheetPayload]:
    workbook = load_workbook(path, data_only=True)
    payloads: List[WorksheetPayload] = []
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        payload = worksheet_to_dataframe(path, ws, max_header_scan=max_header_scan)
        if payload is not None:
            payloads.append(payload)
    return payloads


def main() -> None:
    args = parse_args()
    excel_paths = resolve_files(args.excel_files)
    if not excel_paths:
        raise SystemExit("No Excel files matched the provided arguments.")

    all_payloads: List[WorksheetPayload] = []
    for path in excel_paths:
        print(f"Processing {path} …")
        payloads = process_workbook(path, max_header_scan=args.max_header_scan)
        if not payloads:
            print(f"  No tabular data found in {path}")
            continue
        all_payloads.extend(payloads)

    if not all_payloads:
        raise SystemExit("No usable worksheets found across provided files.")

    write_payloads_to_db(all_payloads, args.db)
    print(f"Wrote {len(all_payloads)} tables to {args.db}")


if __name__ == "__main__":
    main()
